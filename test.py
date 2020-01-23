from tempfile import NamedTemporaryFile

from openpyxl import Workbook
from pyxform import create_survey_from_xls
from lxml import etree as ET


class XLSForm:
    def __init__(self, survey_headers=[]):
        self.wb = Workbook()

        self.survey = self.wb.active
        self.survey.title = 'survey'
        self.choices = self.wb.create_sheet(title='choices')
        self.settings = self.wb.create_sheet(title='settings')

        self.survey.append(['type', 'name', 'label', *survey_headers])
        self.choices.append(['list_name', 'name', 'label'])

        self.survey_headers = survey_headers

    def add_options(self, list_name, options):
        for option in options:
            name, label = option['name'], option['label']
            self.choices.append([list_name, name, label])

    def add_question(self, question_type, name, label, **kwargs):
        extra_fields = [
            '' if header not in kwargs else kwargs[header]
            for header in self.survey_headers
        ]

        self.survey.append([
            question_type,
            name,
            label,
            *extra_fields,
        ])

    def create_xform(self, name='Test form'):
        with NamedTemporaryFile(suffix='.xlsx') as tmp:
            self.wb.save(tmp.name)
            tmp.seek(0)
            survey = create_survey_from_xls(tmp)
            survey.title = name
        return survey.to_xml()

    def create_enketo_form(self, name='Test form'):
        tree = ET.fromstring(self.create_xform(name))

        form_xslt = ET.parse('openrosa2html5form.xsl')
        model_xslt = ET.parse('openrosa2xmlmodel.xsl')

        form_transform = ET.XSLT(form_xslt)
        model_transform = ET.XSLT(model_xslt)

        form = form_transform(tree)
        model = model_transform(tree)

        return {
            'form': ET.tostring(form.getroot()[0]).decode(),
            'model': ET.tostring(model.getroot()[0]).decode(),
        }


# By default, you can provide question_type, name and label for each question.
# However, you can add more fields such as required, default and constraint
# as follows.
form = XLSForm(['required', 'default', 'constraint'])

form.add_question('text', 'name', 'Name')
form.add_question('integer', 'age', 'Age', constraint='. >= 3')
print(form.create_enketo_form())


# See here for question types and other possible fields.
# http://xlsform.org/en/
